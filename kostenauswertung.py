"""
Swisspor Kostenauswertung
=========================
Input:  - Mengenauswertung.xlsx  (IFC-Export aus Solibri: SystemID + Fläche)
        - System_0X_*.csv        (Schichtaufbau + Einheitspreise)
Output: - Kostenauswertung_Swisspor.xlsx

Logik:
  Für jeden Gesamtaufbau (SystemID) aus dem Modell:
    → alle Schichten dieses Systems aus CSV laden
    → Fläche × Einheitspreis = Positionskosten
    → Sonderschichten (Absturzsicherung) → Stk-Preis statt m²
"""

import os, glob
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Pfade ──────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
MENGEN_FILE  = os.path.join(BASE, "mengen_input.xlsx")
CSV_PATTERN  = os.path.join(BASE, "System_*.csv")
OUTPUT_FILE  = os.path.join(BASE, "Kostenauswertung_Swisspor.xlsx")

# ── Styles ─────────────────────────────────────────────────────────────────
BLUE_D  = "1F4E79"
BLUE_M  = "2E75B6"
BLUE_L  = "D5E8F0"
GREY    = "F2F2F2"
WHITE   = "FFFFFF"
GREEN   = "E2EFDA"

def thin_border():
    s = Side(style='thin', color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def style(cell, bold=False, bg=WHITE, fg="000000", align="left",
          size=10, num_fmt=None, wrap=False):
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = thin_border()
    if num_fmt:
        cell.number_format = num_fmt

# ══════════════════════════════════════════════════════════════════════════
# 1. Mengenauswertung laden (IFC-Export)
# ══════════════════════════════════════════════════════════════════════════
print("📂 Lade Mengenauswertung...")
df_mengen = pd.read_excel(MENGEN_FILE, sheet_name=0)
df_mengen.columns = [c.strip() for c in df_mengen.columns]

# Normalisiere Spaltennamen flexibel
col_map = {}
for c in df_mengen.columns:
    cl = c.lower()
    if "systemid" in cl or "system-id" in cl or "system_id" in cl:
        col_map[c] = "SystemID"
    elif "function" in cl:
        col_map[c] = "Function"
    elif "fläche" in cl or "flache" in cl or "area" in cl:
        col_map[c] = "Flaeche"
    elif "volumen" in cl or "volume" in cl:
        col_map[c] = "Volumen"
    elif "dicke" in cl or "thickness" in cl:
        col_map[c] = "Dicke"

df_mengen = df_mengen.rename(columns=col_map)
print(f"   → {len(df_mengen)} Zeilen geladen")
print(f"   → Spalten: {list(df_mengen.columns)}")

# Nur Gesamtaufbau-Zeilen (SystemID gesetzt, Function = Gesamtaufbau oder leer)
df_gesamt = df_mengen[
    df_mengen["SystemID"].notna() &
    (df_mengen["SystemID"].astype(str).str.strip() != "") &
    (
        df_mengen.get("Function", pd.Series(dtype=str))
        .astype(str).str.contains("gesamt|00_|total", case=False, na=True) |
        df_mengen.get("Function", pd.Series(dtype=str))
        .isna()
    )
].copy()

# Falls keine Gesamtaufbau-Zeilen: alle Zeilen mit SystemID nehmen
if len(df_gesamt) == 0:
    df_gesamt = df_mengen[
        df_mengen["SystemID"].notna() &
        (df_mengen["SystemID"].astype(str).str.strip() != "")
    ].drop_duplicates(subset=["SystemID"]).copy()

# Fläche pro System als Dict
flaeche_pro_system = {}
for _, row in df_gesamt.iterrows():
    sid = str(row["SystemID"]).strip().zfill(2)
    flaeche = float(row.get("Flaeche", 0) or 0)
    if flaeche > 0:
        flaeche_pro_system[sid] = flaeche

print(f"   → Systeme mit Fläche: {flaeche_pro_system}")

# ══════════════════════════════════════════════════════════════════════════
# 2. System-CSVs laden
# ══════════════════════════════════════════════════════════════════════════
print("\n📂 Lade System-CSVs...")
csv_files = sorted(glob.glob(CSV_PATTERN))
df_all = pd.concat([pd.read_csv(f) for f in csv_files], ignore_index=True)
df_all["SystemID"] = df_all["SystemID"].astype(str).str.zfill(2)
print(f"   → {len(df_all)} Schichten aus {len(csv_files)} Dateien geladen")

# ══════════════════════════════════════════════════════════════════════════
# 3. Kostenberechnung
# ══════════════════════════════════════════════════════════════════════════
print("\n⚙️  Berechne Kosten...")
rows = []

for system_id, flaeche in sorted(flaeche_pro_system.items()):
    schichten = df_all[df_all["SystemID"] == system_id].copy()
    
    if len(schichten) == 0:
        print(f"   ⚠️  Keine Schichten für System {system_id}")
        continue

    system_total = 0.0
    
    for _, s in schichten.iterrows():
        preis     = float(s.get("Einheitspreis", 0) or 0)
        preisbasis = str(s.get("Preisbasis", "m²")).strip()
        funktion  = str(s.get("Function", "")).strip()
        
        # Sonderfall Absturzsicherung → Stückpreis, Menge = 1
        if preisbasis == "Stk" or "absturz" in funktion.lower():
            menge        = 1
            einheit      = "Stk"
            poskosten    = preis * menge
        else:
            menge        = flaeche
            einheit      = "m²"
            poskosten    = preis * menge

        system_total += poskosten

        rows.append({
            "SystemID":        system_id,
            "Systemname":      f"System {system_id}",
            "Function":        funktion,
            "Produkt":         str(s.get("Produktbezeichnung", "")),
            "Artikelnummer":   str(s.get("Artikelnummer", "")),
            "Kostencode":      str(s.get("Kostencode", "")),
            "Preisbasis":      einheit,
            "Menge":           round(menge, 2),
            "Einheitspreis":   preis,
            "Positionskosten": round(poskosten, 2),
            "_system_total":   0,  # placeholder
        })

    # Mark system total on last row
    if rows:
        rows[-1]["_system_total"] = round(system_total, 2)

print(f"   → {len(rows)} Positionen berechnet")

# ══════════════════════════════════════════════════════════════════════════
# 4. Excel erstellen
# ══════════════════════════════════════════════════════════════════════════
print("\n📊 Erstelle Excel...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kostenauswertung"

# Titel
ws.merge_cells("A1:J1")
c = ws.cell(row=1, column=1,
    value="Swisspor – Kostenauswertung Systemaufbauten  |  exkl. MWST  |  Preisstand 2026")
c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
c.fill      = PatternFill("solid", fgColor=BLUE_D)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Flächen-Übersicht
ws.merge_cells("A2:J2")
flaechen_txt = "  |  ".join([f"System {k}: {v:.2f} m²" for k,v in sorted(flaeche_pro_system.items())])
c2 = ws.cell(row=2, column=1, value=f"Flächen aus IFC-Modell:  {flaechen_txt}")
c2.font      = Font(name="Arial", size=10, color=WHITE)
c2.fill      = PatternFill("solid", fgColor=BLUE_M)
c2.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

# Header
headers = [
    "System", "Systemname", "Schichtfunktion", "Produkt",
    "Art.-Nr.", "BKP-Code", "Einheit", "Menge",
    "EP CHF", "Positionskosten CHF"
]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    style(c, bold=True, bg=BLUE_M, fg=WHITE, align="center", size=10, wrap=True)
ws.row_dimensions[3].height = 28

# Daten
current_system = None
row_num = 4
grand_total = 0.0

for entry in rows:
    sid = entry["SystemID"]
    
    # System-Trennlinie
    if sid != current_system:
        if current_system is not None:
            # Leere Trennzeile
            for col in range(1, 11):
                c = ws.cell(row=row_num, column=col, value="")
                style(c, bg=BLUE_L)
            ws.row_dimensions[row_num].height = 6
            row_num += 1
        current_system = sid

    # Zebra
    bg = GREY if row_num % 2 == 0 else WHITE
    
    vals = [
        entry["SystemID"],
        entry["Systemname"],
        entry["Function"],
        entry["Produkt"],
        entry["Artikelnummer"],
        entry["Kostencode"],
        entry["Preisbasis"],
        entry["Menge"],
        entry["Einheitspreis"],
        entry["Positionskosten"],
    ]
    aligns = ["center","left","left","left","center","center","center","right","right","right"]
    
    for i, (val, aln) in enumerate(zip(vals, aligns), 1):
        c = ws.cell(row=row_num, column=i, value=val)
        if i == 8:
            style(c, bg=bg, align=aln, num_fmt='#,##0.00')
        elif i == 9:
            style(c, bg=bg, align=aln, num_fmt='#,##0.00 "CHF"')
        elif i == 10:
            bold = entry["Positionskosten"] == 0
            style(c, bg=GREEN if entry["Positionskosten"] > 0 else bg,
                  align=aln, num_fmt='#,##0.00 "CHF"', bold=False)
        else:
            style(c, bg=bg, align=aln)
    ws.row_dimensions[row_num].height = 16
    row_num += 1
    
    # System-Total nach letzter Schicht
    if entry["_system_total"] > 0:
        grand_total += entry["_system_total"]
        ws.merge_cells(f"A{row_num}:I{row_num}")
        c = ws.cell(row=row_num, column=1,
            value=f"  Subtotal System {sid}")
        style(c, bold=True, bg=BLUE_M, fg=WHITE, align="left", size=10)
        c2 = ws.cell(row=row_num, column=10,
            value=entry["_system_total"])
        style(c2, bold=True, bg=BLUE_M, fg=WHITE, align="right",
              size=10, num_fmt='#,##0.00 "CHF"')
        ws.row_dimensions[row_num].height = 20
        row_num += 1

# Leerzeile
for col in range(1, 11):
    c = ws.cell(row=row_num, column=col, value="")
    style(c, bg=WHITE)
row_num += 1

# Grand Total
ws.merge_cells(f"A{row_num}:I{row_num}")
c = ws.cell(row=row_num, column=1, value="TOTAL ALLE SYSTEME  (exkl. MWST)")
style(c, bold=True, bg=BLUE_D, fg=WHITE, align="right", size=12)
c2 = ws.cell(row=row_num, column=10, value=grand_total)
style(c2, bold=True, bg=BLUE_D, fg=WHITE, align="right",
      size=12, num_fmt='#,##0.00 "CHF"')
ws.row_dimensions[row_num].height = 26

# MWST
row_num += 1
ws.merge_cells(f"A{row_num}:I{row_num}")
c = ws.cell(row=row_num, column=1, value="MWST 8.1%")
style(c, bold=True, bg=BLUE_L, align="right", size=11)
c2 = ws.cell(row=row_num, column=10, value=round(grand_total * 0.081, 2))
style(c2, bold=True, bg=BLUE_L, align="right",
      size=11, num_fmt='#,##0.00 "CHF"')
ws.row_dimensions[row_num].height = 22

# Total inkl MWST
row_num += 1
ws.merge_cells(f"A{row_num}:I{row_num}")
c = ws.cell(row=row_num, column=1, value="TOTAL inkl. MWST")
style(c, bold=True, bg=BLUE_D, fg=WHITE, align="right", size=12)
c2 = ws.cell(row=row_num, column=10, value=round(grand_total * 1.081, 2))
style(c2, bold=True, bg=BLUE_D, fg=WHITE, align="right",
      size=12, num_fmt='#,##0.00 "CHF"')
ws.row_dimensions[row_num].height = 26

# Spaltenbreiten
widths = [10, 14, 26, 34, 14, 14, 10, 10, 14, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A4"

# ── Sheet 2: Mengen-Mapping ─────────────────────────────────────────────
ws2 = wb.create_sheet("Mengen_Mapping")
ws2.merge_cells("A1:E1")
c = ws2.cell(row=1, column=1,
    value="Flächengrundlagen aus IFC-Modell (Gesamtaufbau pro System)")
c.font = Font(name="Arial", bold=True, size=12, color=WHITE)
c.fill = PatternFill("solid", fgColor=BLUE_D)
c.alignment = Alignment(horizontal="center", vertical="center")

for i, h in enumerate(["SystemID","Fläche m²","Anzahl Schichten","Subtotal CHF","Quelle"], 1):
    c = ws2.cell(row=2, column=i, value=h)
    style(c, bold=True, bg=BLUE_M, fg=WHITE, align="center")

for r, (sid, fl) in enumerate(sorted(flaeche_pro_system.items()), 3):
    schichten_n = len(df_all[df_all["SystemID"] == sid])
    subtotal = sum(e["_system_total"] for e in rows
                   if e["SystemID"] == sid and e["_system_total"] > 0)
    bg = GREY if r % 2 == 0 else WHITE
    for i, val in enumerate([sid, fl, schichten_n, subtotal, "Solibri ITO Export"], 1):
        c = ws2.cell(row=r, column=i, value=val)
        fmt = '#,##0.00 "CHF"' if i == 4 else ('#,##0.00' if i == 2 else None)
        style(c, bg=bg, align="center" if i != 5 else "left", num_fmt=fmt)

for i, w in enumerate([12, 14, 18, 20, 24], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save(OUTPUT_FILE)
print(f"\n✅ Fertig: {OUTPUT_FILE}")
print(f"   Total exkl. MWST: CHF {grand_total:,.2f}")
print(f"   Total inkl. MWST: CHF {grand_total*1.081:,.2f}")
EOF
