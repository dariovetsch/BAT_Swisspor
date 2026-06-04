import streamlit as st
import swisspor_mapping
import swisspor_kosten
import io
import openpyxl

st.set_page_config(
    page_title="swissporBIM",
    page_icon="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 60 40'><text y='28' font-size='18' font-family='Helvetica Neue,sans-serif' font-weight='700' fill='%231a3a5c'>sw</text><circle cx='44' cy='22' r='5' fill='%231a3a5c'/><circle cx='52' cy='22' r='5' fill='%23f0a500'/><circle cx='60' cy='22' r='5' fill='%23e63946'/></svg>",
    layout="centered",
    initial_sidebar_state="collapsed"
)


@st.cache_data(show_spinner=False)
def _load_dritt_defaults(file_bytes: bytes) -> dict:
    """Liest EP Material + EP Einbau je CostLabel aus dem Sheet 'Lookup_Drittprodukte'.
    Dient als Vorbelegung der Eingabefelder – Preise werden einmal im Sheet gepflegt."""
    out = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        dk = next((s for s in wb.sheetnames if "drittprodukt" in s.lower()), None)
        if dk:
            for row in wb[dk].iter_rows(min_row=3, values_only=True):
                if not row or row[0] is None:
                    continue
                label = str(row[0]).strip()
                try:    preis  = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
                except (TypeError, ValueError): preis = 0.0
                try:    einbau = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
                except (TypeError, ValueError): einbau = 0.0
                out[label] = {"preis": preis, "einbau": einbau}
    except Exception:
        pass
    return out


st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Source+Sans+3:wght@300;400;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">

<style>
:root {
    --sw-navy:   #1a3a5c;
    --sw-red:    #e63946;
    --sw-orange: #f0a500;
    --sw-bg:     #f0f2f5;
    --sw-white:  #ffffff;
    --sw-border: #e0e0e0;
    --sw-text:   #1a1a1a;
    --sw-muted:  #6b7280;
    --font:      'Source Sans 3', 'Helvetica Neue', Helvetica, Arial, sans-serif;
    --mono:      'DM Mono', monospace;
}

#MainMenu, footer, header, .stDeployButton,
[data-testid="collapsedControl"] { display: none !important; }
.block-container { padding: 0 !important; max-width: 900px !important; margin: 0 auto !important; }
section[data-testid="stSidebar"] { display: none !important; }

html, body, .stApp {
    background: var(--sw-bg) !important;
    font-family: var(--font) !important;
    color: var(--sw-text) !important;
}

/* ── Topnav ── */
.sw-nav {
    background: var(--sw-white);
    border-bottom: 3px solid var(--sw-red);
    padding: 0 3rem;
    height: 60px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    position: sticky;
    top: 0;
    z-index: 200;
}
.sw-logo { display: flex; align-items: center; }
.sw-logo-text { font-size: 1.35rem; font-weight: 700; color: var(--sw-navy);
    letter-spacing: -0.01em; line-height: 1; }
.sw-logo-dots { display: flex; gap: 4px; margin-left: 6px; margin-top: 2px; }
.sw-logo-dots span { width: 9px; height: 9px; border-radius: 50%; display: inline-block; }
.sw-nav-title { font-size: 0.95rem; color: var(--sw-navy); font-weight: 400; }
.sw-nav-right { font-size: 0.78rem; color: var(--sw-muted); font-family: var(--mono);
    letter-spacing: 0.05em; }

/* ── Page wrapper ── */
.sw-page {
    padding: 1.5rem 0 4rem 0;
    max-width: 900px;
    margin: 0 auto;
}

/* ── Tabs ── */
[data-testid="stTabs"] [role="tablist"] {
    background: transparent !important; border-bottom: 2px solid var(--sw-border) !important;
    gap: 0 !important; padding: 0 !important; }
[data-testid="stTabs"] [role="tab"] {
    font-family: var(--font) !important; font-size: 0.78rem !important;
    font-weight: 700 !important; letter-spacing: 0.1em !important;
    text-transform: uppercase !important; color: var(--sw-muted) !important;
    padding: 0.7rem 1.6rem !important; border: none !important;
    border-bottom: 2px solid transparent !important; border-radius: 0 !important;
    background: transparent !important; margin-bottom: -2px !important; }
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    color: var(--sw-navy) !important; border-bottom-color: var(--sw-red) !important;
    background: transparent !important; }
[data-testid="stTabsContent"] {
    padding-top: 1.5rem !important; border: none !important; }

/* ── Cards ── */
.sw-card {
    background: var(--sw-white);
    border-radius: 6px;
    padding: 1.75rem 2rem;
    margin-bottom: 1.25rem;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}

/* ── Section headers ── */
.sw-section-header { display: flex; align-items: center; gap: 12px; margin-bottom: 1rem; }
.sw-section-num { width: 26px; height: 26px; border-radius: 50%;
    background: var(--sw-navy); color: white; font-size: 0.72rem; font-weight: 700;
    display: flex; align-items: center; justify-content: center; flex-shrink: 0; }
.sw-section-title { font-size: 0.78rem; font-weight: 700; letter-spacing: 0.12em;
    text-transform: uppercase; color: var(--sw-navy); }
.sw-upload-hint { font-size: 0.82rem; color: var(--sw-muted);
    margin-bottom: 1rem; line-height: 1.4; }

.sw-file-badge { display: inline-flex; align-items: center; gap: 5px;
    color: var(--sw-navy); border: 1px solid var(--sw-navy);
    font-size: 0.75rem; font-weight: 600; padding: 3px 10px;
    border-radius: 2px; margin: 3px 5px 3px 0; }

/* ── Buttons ── */
.stButton > button {
    background: var(--sw-white) !important; color: var(--sw-red) !important;
    border: 1.5px solid var(--sw-red) !important; border-radius: 2px !important;
    font-family: var(--font) !important; font-size: 0.88rem !important;
    font-weight: 600 !important; letter-spacing: 0.02em !important;
    padding: 0.55rem 1.8rem !important; transition: all 0.15s !important; }
.stButton > button:hover { background: var(--sw-red) !important; color: white !important; }
.stButton > button:disabled { opacity: 0.35 !important; }

/* ── Stats strip ── */
.sw-stats { display: grid; grid-template-columns: repeat(4, 1fr);
    gap: 1px; background: var(--sw-border);
    border: 1px solid var(--sw-border); border-radius: 6px;
    overflow: hidden; margin: 1.5rem 0; }
.sw-stat { background: var(--sw-white); padding: 1.2rem 1.5rem; text-align: center; }
.sw-stat-num { font-size: 2rem; font-weight: 700; color: var(--sw-navy);
    line-height: 1; margin-bottom: 0.3rem; }
.sw-stat-num.red   { color: var(--sw-red); }
.sw-stat-num.green { color: #2d7d46; }
.sw-stat-num.warn  { color: #c47f00; }
.sw-stat-num.chf   { font-size: 1.2rem; color: var(--sw-navy); }
.sw-stat-label { font-size: 0.68rem; font-weight: 700; letter-spacing: 0.1em;
    text-transform: uppercase; color: var(--sw-muted); }

/* ── Log ── */
.sw-log { background: #f9f9f9; border: 1px solid var(--sw-border);
    border-left: 3px solid var(--sw-navy); border-radius: 2px;
    font-family: var(--mono); font-size: 0.74rem; color: #444;
    padding: 1rem 1.2rem; line-height: 1.75; max-height: 280px;
    overflow-y: auto; white-space: pre-wrap; }
.sw-log::-webkit-scrollbar { width: 3px; }
.sw-log::-webkit-scrollbar-thumb { background: var(--sw-border); }

.stDownloadButton > button {
    background: var(--sw-red) !important; color: white !important;
    border: none !important; border-radius: 2px !important;
    font-family: var(--font) !important; font-size: 0.88rem !important;
    font-weight: 600 !important; letter-spacing: 0.04em !important;
    padding: 0.6rem 2rem !important; transition: opacity 0.15s !important; }
.stDownloadButton > button:hover { opacity: 0.85 !important; }

.sw-success-bar { background: var(--sw-navy); color: white;
    padding: 0.65rem 1.4rem; font-size: 0.82rem; font-weight: 600;
    letter-spacing: 0.04em; border-radius: 4px; margin-bottom: 1.2rem;
    display: flex; align-items: center; gap: 8px; }
.sw-success-bar::before { content: ''; width: 8px; height: 8px;
    border-radius: 50%; background: #5cb85c; flex-shrink: 0; }

[data-testid="stFileUploader"], [data-testid="stFileUploader"] section,
[data-testid="stFileUploader"] > div,
[data-testid="stFileUploaderDropzone"] {
    background: #fafafa !important; background-color: #fafafa !important;
    border: 1.5px dashed #c8d0da !important; border-radius: 3px !important;
    color: var(--sw-text) !important; }
[data-testid="stFileUploaderDropzone"]:hover {
    border-color: var(--sw-navy) !important; background: #f4f6f9 !important; }
[data-testid="stFileUploaderDropzone"] svg { color: var(--sw-navy) !important;
    fill: var(--sw-navy) !important; }
[data-testid="stFileUploaderDropzoneInstructions"] > div > span,
[data-testid="stFileUploaderDropzoneInstructions"] span {
    color: var(--sw-navy) !important; font-family: var(--font) !important;
    font-size: 0.88rem !important; font-weight: 600 !important; }
[data-testid="stFileUploaderDropzoneInstructions"] > div > small,
[data-testid="stFileUploaderDropzoneInstructions"] small {
    color: var(--sw-muted) !important; font-size: 0.75rem !important; }
[data-testid="stFileUploaderDropzone"] button {
    background: white !important; color: var(--sw-navy) !important;
    border: 1.5px solid var(--sw-navy) !important; border-radius: 2px !important;
    font-family: var(--font) !important; font-size: 0.8rem !important;
    font-weight: 600 !important; }
[data-testid="stFileUploaderDropzone"] button:hover {
    background: var(--sw-navy) !important; color: white !important; }
[data-testid="stFileUploaderFile"], [data-testid="uploadedFileData"] {
    background: white !important; border: 1px solid var(--sw-border) !important;
    border-radius: 2px !important; color: var(--sw-text) !important; }
[data-testid="stFileUploaderFileName"] { color: var(--sw-navy) !important; font-weight: 600 !important; }
[data-testid="stFileUploaderFileSize"] { color: var(--sw-muted) !important;
    font-family: var(--mono) !important; font-size: 0.72rem !important; }

.stSpinner > div > div { border-top-color: var(--sw-red) !important; }
.sw-divider { border: none; border-top: 1px solid var(--sw-border); margin: 1.2rem 0; }
.sw-chart-title { font-size: 0.78rem; font-weight: 700; letter-spacing: 0.1em;
    text-transform: uppercase; color: var(--sw-navy); margin-bottom: 0.75rem; }
.sw-footer { text-align: center; font-size: 0.72rem; color: var(--sw-muted);
    padding: 2rem 0 3rem; letter-spacing: 0.04em; font-family: var(--mono); }

/* ── Abschnitt-Trennlinie ── */
.sw-section-bar {
    display: flex; align-items: center; gap: 12px;
    margin: 2.5rem 0 1rem;
}
.sw-section-bar-line {
    width: 3px; height: 22px; border-radius: 2px; flex-shrink: 0;
}
.sw-section-bar-title {
    font-size: 0.78rem; font-weight: 700; letter-spacing: 0.12em;
    text-transform: uppercase; color: var(--sw-navy);
}
.sw-section-bar-sub {
    font-size: 0.82rem; color: #6b7280; margin: 0.15rem 0 1rem 15px; line-height: 1.5;
}
</style>
""", unsafe_allow_html=True)

# ─── TOPNAV ───────────────────────────────────────────────────────────────────
st.markdown("""
<nav class="sw-nav">
    <div class="sw-logo">
        <span class="sw-logo-text">swisspor</span>
        <div class="sw-logo-dots">
            <span style="background:#1a3a5c"></span>
            <span style="background:#f0a500"></span>
            <span style="background:#e63946"></span>
        </div>
    </div>
    <span class="sw-nav-title">swissporBIM · IFC-Attribuierung &amp; Kostenermittlung</span>
    <span class="sw-nav-right">DC_BAT_FS26 · v5.0</span>
</nav>
""", unsafe_allow_html=True)

st.markdown('<div class="sw-page">', unsafe_allow_html=True)

tab1, tab2 = st.tabs([
    "UC-01  ·  IFC-Mapping",
    "UC-02  ·  Mengen & Kosten"
])

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 – MAPPING
# ════════════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="sw-card">', unsafe_allow_html=True)
    st.markdown("""
    <div class="sw-section-header">
        <div class="sw-section-num">1</div>
        <span class="sw-section-title">IFC-Modell hochladen</span>
    </div>
    <p class="sw-upload-hint">ArchiCAD / Revit Export mit <code>Pset_Swisspor_Identifikation</code>.</p>
    """, unsafe_allow_html=True)
    ifc_file = st.file_uploader("IFC", type=["ifc"], label_visibility="collapsed", key="ifc_map")
    if ifc_file:
        size = len(ifc_file.getvalue()) / 1024 / 1024
        st.markdown(f'<div class="sw-file-badge">↑ {ifc_file.name} · {size:.1f} MB</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sw-card">', unsafe_allow_html=True)
    st.markdown("""
    <div class="sw-section-header">
        <div class="sw-section-num">2</div>
        <span class="sw-section-title">System-Kataloge hochladen</span>
    </div>
    <p class="sw-upload-hint">System_01_GREENROOF.csv, System_02_KIES_SAFSYS.csv, …</p>
    """, unsafe_allow_html=True)
    csv_files = st.file_uploader("CSVs", type=["csv"], accept_multiple_files=True,
                                  label_visibility="collapsed", key="csv_map")
    if csv_files:
        st.markdown("".join(f'<div class="sw-file-badge">↑ {f.name}</div>' for f in csv_files),
                    unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sw-card">', unsafe_allow_html=True)
    st.markdown("""
    <div class="sw-section-header">
        <div class="sw-section-num">3</div>
        <span class="sw-section-title">Mapping ausführen</span>
    </div>
    """, unsafe_allow_html=True)
    if not (ifc_file and csv_files):
        st.markdown('<p class="sw-upload-hint">Bitte IFC-Modell und CSV-Kataloge hochladen.</p>', unsafe_allow_html=True)
    col_btn, _ = st.columns([1, 3])
    with col_btn:
        run_map = st.button("Mapping starten →", disabled=not (ifc_file and csv_files),
                             use_container_width=True, key="btn_map")
    st.markdown('</div>', unsafe_allow_html=True)

    if run_map and ifc_file and csv_files:
        with st.spinner("Mapping läuft …"):
            try:
                results = swisspor_mapping.process_mapping(ifc_file, csv_files)
            except Exception as e:
                import traceback
                st.error(f"Fehler: {e}")
                st.code(traceback.format_exc())
                st.stop()

        df = results["summary_table"]
        def _n(label):
            row = df[df["Kategorie"].str.contains(label, na=False)]
            return int(row["Anzahl"].values[0]) if len(row) else 0

        st.markdown(f'<div class="sw-success-bar">Mapping abgeschlossen — {len(csv_files)} Katalog(e) verarbeitet</div>',
                    unsafe_allow_html=True)
        st.markdown(f"""
        <div class="sw-stats">
            <div class="sw-stat"><div class="sw-stat-num green">{_n("direkt")}</div><div class="sw-stat-label">Direkt gemappt</div></div>
            <div class="sw-stat"><div class="sw-stat-num">{_n("Gesamtaufbau")}</div><div class="sw-stat-label">Gesamtaufbau</div></div>
            <div class="sw-stat"><div class="sw-stat-num warn">{_n("Kein CSV")}</div><div class="sw-stat-label">Kein CSV</div></div>
            <div class="sw-stat"><div class="sw-stat-num" style="color:#999">{_n("bersprungen")}</div><div class="sw-stat-label">Übersprungen</div></div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="sw-card">', unsafe_allow_html=True)
        st.markdown('<div class="sw-section-header"><div class="sw-section-num" style="background:#555">4</div><span class="sw-section-title">Protokoll &amp; Download</span></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="sw-log">{results["log"]}</div>', unsafe_allow_html=True)
        st.markdown('<hr class="sw-divider">', unsafe_allow_html=True)
        out_name = ifc_file.name.replace(".ifc", "_enriched.ifc")
        st.download_button(f"↓  {out_name} herunterladen", data=results["ifc_bytes"],
                           file_name=out_name, mime="application/octet-stream")
        st.markdown('</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 – MENGEN & KOSTEN
# ════════════════════════════════════════════════════════════════════════════
with tab2:

    st.markdown('<div class="sw-card">', unsafe_allow_html=True)
    st.markdown("""
    <div class="sw-section-header">
        <div class="sw-section-num">1</div>
        <span class="sw-section-title">Enriched IFC hochladen</span>
    </div>
    <p class="sw-upload-hint">Das angereicherte IFC aus UC-01 – enthält alle Pset_Swisspor_* Schichten.</p>
    """, unsafe_allow_html=True)
    ifc_enriched = st.file_uploader("Enriched IFC", type=["ifc"],
                                     label_visibility="collapsed", key="ifc_kost")
    if ifc_enriched:
        size = len(ifc_enriched.getvalue()) / 1024 / 1024
        st.markdown(f'<div class="sw-file-badge">↑ {ifc_enriched.name} · {size:.1f} MB</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sw-card">', unsafe_allow_html=True)
    st.markdown("""
    <div class="sw-section-header">
        <div class="sw-section-num">2</div>
        <span class="sw-section-title">Preisdatenbank hochladen</span>
    </div>
    <p class="sw-upload-hint">
        Swisspor_Preisdatenbank_2026.xlsx – Sheets <code>Lookup_Kostenauswertung</code>
        (Art.-Nr. + EP_Einbau) und <code>Lookup_Drittprodukte</code> (CostLabel → CHF).
    </p>
    """, unsafe_allow_html=True)
    preisdb_file = st.file_uploader("Preisdatenbank", type=["xlsx"],
                                     label_visibility="collapsed", key="preisdb")
    if preisdb_file:
        st.markdown(f'<div class="sw-file-badge">↑ {preisdb_file.name}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Drittprodukte Materialpreise ──────────────────────────────────────────
    DRITT_ROWS = [
        ("Schutzschicht Kies",      "Rundkies 50 mm",           11.0),
        ("Splitschicht",            "Splittbett 50 mm",          15.0),
        ("Vegetationstragschicht",  "Extensivsubstrat 120 mm",   20.0),
        ("Belag geschlossen",       "Beton-/Natursteinplatten",  90.0),
        ("Belag offen",             "Holz-/WPC-Belag",           80.0),
        ("Brandschutzschicht",      "Brandschutzvlies",           5.5),
    ]
    if preisdb_file:
        st.markdown('<div class="sw-card">', unsafe_allow_html=True)
        st.markdown("""
        <div class="sw-section-header">
            <div class="sw-section-num">3</div>
            <span class="sw-section-title">Drittprodukte — Materialpreise erfassen</span>
        </div>
        <p class="sw-upload-hint">
            Diese Schichten liefert swisspor nicht. Material- und Einbaukosten kommen aus der
            Preisdatenbank (Blatt <code>Lookup_Drittprodukte</code>) und sind hier bereits
            vorausgefüllt — du kannst sie projektspezifisch überschreiben.
        </p>
        """, unsafe_allow_html=True)

        col_h = st.columns([3, 1.4, 1.4, 1.6])
        for c, t in zip(col_h, ["Schicht / Produkt", "EP Material CHF/m²", "EP Einbau CHF/m²", ""]):
            c.markdown(f"<div style='font-size:0.65rem;font-weight:700;letter-spacing:0.1em;"
                       f"text-transform:uppercase;color:#9ca3af;padding:0.25rem 0;'>{t}</div>",
                       unsafe_allow_html=True)

        dritt_defaults = _load_dritt_defaults(preisdb_file.getvalue())

        dritt_inputs = {}
        for label, bez, ein_default in DRITT_ROWS:
            dflt        = dritt_defaults.get(label, {})
            mat_default = float(dflt.get("preis", 0.0))
            ein_used    = float(dflt.get("einbau", ein_default))
            row_c = st.columns([3, 1.4, 1.4, 1.6])
            with row_c[0]:
                st.markdown(f"<div style='font-size:0.83rem;padding:0.4rem 0;'>"
                            f"<b>{label}</b>"
                            f"<span style='font-size:0.72rem;color:#9ca3af;margin-left:8px;'>{bez}</span>"
                            f"</div>", unsafe_allow_html=True)
            with row_c[1]:
                mat_val = st.number_input("", min_value=0.0, max_value=9999.0,
                                          value=mat_default, step=0.5, format="%.2f",
                                          key=f"dm_{label}", label_visibility="collapsed")
            with row_c[2]:
                st.markdown(f"<div style='font-size:0.82rem;color:#9ca3af;"
                            f"font-family:monospace;padding:0.6rem 0.2rem;'>{ein_used:.2f}</div>",
                            unsafe_allow_html=True)
            with row_c[3]:
                if mat_val > 0:
                    st.markdown("<div style='font-size:0.72rem;color:#2d7d46;"
                                "padding:0.6rem 0;font-weight:600;'>✓ erfasst</div>",
                                unsafe_allow_html=True)
                else:
                    st.markdown("<div style='font-size:0.72rem;color:#9ca3af;"
                                "padding:0.6rem 0;'>— offen</div>", unsafe_allow_html=True)
            dritt_inputs[label] = {"preis": mat_val, "einbau": ein_used}

        st.markdown("<div style='font-size:0.72rem;color:#9ca3af;margin-top:0.25rem;'>"
                    "Einbaukosten: Richtwerte NPK 364 (CRB 2025), Mittellohn CHF 114.88/h (suissetec 2025). "
                    "Alle Preise exkl. MWST.</div>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        dritt_inputs = {label: {"preis": 0.0, "einbau": ein} for label, _, ein in DRITT_ROWS}

    st.markdown('<div class="sw-card">', unsafe_allow_html=True)
    st.markdown("""
    <div class="sw-section-header">
        <div class="sw-section-num">4</div>
        <span class="sw-section-title">Auswertung berechnen</span>
    </div>
    """, unsafe_allow_html=True)
    if not (ifc_enriched and preisdb_file):
        st.markdown('<p class="sw-upload-hint">Bitte enriched IFC und Preisdatenbank hochladen.</p>', unsafe_allow_html=True)
    col_btn2, _ = st.columns([1, 3])
    with col_btn2:
        run_kost = st.button("Kosten berechnen →", disabled=not (ifc_enriched and preisdb_file),
                              use_container_width=True, key="btn_kost")
    st.markdown('</div>', unsafe_allow_html=True)

    if run_kost and ifc_enriched and preisdb_file:
        with st.spinner("Auswertung läuft …"):
            try:
                results_k = swisspor_kosten.process_kosten(ifc_enriched, preisdb_file, dritt_inputs)
            except Exception as e:
                import traceback
                st.error(f"Fehler: {e}")
                st.code(traceback.format_exc())
                st.stop()

        gt     = results_k["grand_total"]
        n_sw   = results_k.get("n_matched", 0)
        n_drit = results_k.get("n_drittprodukt", 0)
        n_sys  = results_k["n_systems"]

        st.markdown('<div class="sw-success-bar">Auswertung abgeschlossen</div>',
                    unsafe_allow_html=True)
        st.markdown(
            "<div style='font-size:0.76rem;color:#92400e;background:#fef9ec;"
            "border-left:3px solid #f0a500;padding:0.55rem 1rem;"
            "border-radius:4px;margin-bottom:0.75rem;line-height:1.5;'>"
            "⚠️ <b>Hinweis:</b> Einbaukosten basieren auf Richtwerten (NPK 364, CRB 2025). "
            "Projektspezifische Zuschläge (Gerüst, Anschlüsse, Randdetails) sind nicht enthalten. "
            "Werte nur für Demonstrationszwecke.</div>",
            unsafe_allow_html=True
        )
        st.markdown(f"""
        <div class="sw-stats">
            <div class="sw-stat"><div class="sw-stat-num green">{n_sw}</div><div class="sw-stat-label">swisspor Pos.</div></div>
            <div class="sw-stat"><div class="sw-stat-num warn">{n_drit}</div><div class="sw-stat-label">Drittprodukte</div></div>
            <div class="sw-stat"><div class="sw-stat-num">{n_sys}</div><div class="sw-stat-label">Systeme</div></div>
            <div class="sw-stat"><div class="sw-stat-num chf">CHF {gt:,.0f}</div><div class="sw-stat-label">Total exkl. MWST</div></div>
        </div>
        """, unsafe_allow_html=True)

        kosten_list = results_k.get("kosten_list", [])
        if kosten_list:
            import pandas as pd
            import plotly.graph_objects as go

            df_k = pd.DataFrame(kosten_list)

            # Funktion-Label säubern (Nummer + Unterstrich entfernen)
            df_k["func_label"] = (df_k["function"]
                                  .str.replace(r"^\d+_", "", regex=True)
                                  .str.strip())

            PALETTE = ["#1a3a5c", "#e63946", "#f0a500", "#2d7d46",
                       "#0077b6", "#8338ec", "#fb8500", "#6b7280"]

            systems = sorted(df_k["system_id"].unique())

            # ────────────────────────────────────────────────────────────────
            # ABSCHNITT A – MENGENAUSWERTUNG
            # ────────────────────────────────────────────────────────────────
            st.markdown("""
            <div class="sw-section-bar">
              <div class="sw-section-bar-line" style="background:#1a3a5c;"></div>
              <span class="sw-section-bar-title">A · Mengenauswertung nach Systemaufbau</span>
            </div>
            <p class="sw-section-bar-sub">
              Alle Schichten je System mit Beschrieb, Fläche, Bestellmengen und Einheitspreis.
            </p>
            """, unsafe_allow_html=True)

            for idx, sid in enumerate(systems):
                sys_df = df_k[df_k["system_id"] == sid].copy().sort_values("function")
                col_accent = PALETTE[idx % len(PALETTE)]
                n_schichten = len(sys_df)
                n_dritt_sys = int(sys_df["is_drittprodukt"].sum())

                base_area_rows = sys_df[sys_df["function"].str.startswith("01_")]
                sys_area = (base_area_rows["flaeche_m2"].sum()
                            if not base_area_rows.empty
                            else sys_df["flaeche_m2"].max())

                dritt_badge = (
                    f'<span style="font-size:0.7rem;background:rgba(255,255,255,0.2);'
                    f'color:white;padding:2px 8px;border-radius:3px;margin-left:8px;'
                    f'font-weight:600;">{n_dritt_sys} Drittprodukt(e)</span>'
                    if n_dritt_sys > 0 else ""
                )

                # ── Alle Zeilen als HTML-Table-Rows bauen ────────────────────
                rows_html = ""
                for i, (_, row) in enumerate(sys_df.iterrows()):
                    bg   = "#fafbfc" if i % 2 == 0 else "white"
                    is_d = row["is_drittprodukt"]
                    prod_c = "#b45309" if is_d else "#1a3a5c"

                    art_display = row["article_nr"] if row["article_nr"] else "—"
                    dritt_tag = (
                        '<span style="font-size:0.6rem;background:#fef3c7;color:#92400e;'
                        'padding:1px 5px;border-radius:2px;margin-left:5px;font-weight:600;'
                        'white-space:nowrap;">Dritt</span>'
                        if is_d else ""
                    )

                    rows_html += f"""<tr style="background:{bg};border-bottom:1px solid #f0f2f5;">
                      <td style="width:26px;padding:0.7rem 0 0.7rem 1.75rem;vertical-align:top;">
                        <div style="width:7px;height:7px;border-radius:50%;background:{col_accent};margin-top:3px;"></div>
                      </td>
                      <td style="width:170px;padding:0.7rem 1rem 0.7rem 0;vertical-align:top;
                          font-size:0.80rem;font-weight:600;color:#374151;
                          font-family:'Source Sans 3',Arial,sans-serif;">
                        {row['func_label']}
                      </td>
                      <td style="padding:0.7rem 1rem 0.7rem 0;vertical-align:top;">
                        <div style="font-size:0.85rem;font-weight:600;color:{prod_c};
                            font-family:'Source Sans 3',Arial,sans-serif;line-height:1.3;">
                          {row['product_name'] or '—'}{dritt_tag}
                        </div>
                        <div style="font-size:0.72rem;color:#9ca3af;font-family:'DM Mono',monospace;
                            margin-top:2px;letter-spacing:0.02em;">
                          {art_display}
                        </div>
                      </td>
                      <td style="width:90px;padding:0.7rem 1rem 0.7rem 0;vertical-align:top;text-align:right;">
                        <div style="font-size:0.88rem;font-weight:600;color:#1a3a5c;
                            font-variant-numeric:tabular-nums;white-space:nowrap;line-height:1.2;">
                          {row['flaeche_m2']:,.2f}
                        </div>
                        <div style="font-size:0.60rem;color:#b0b8c4;margin-top:2px;">m²</div>
                      </td>
                      <td style="width:185px;padding:0.7rem 1rem 0.7rem 0;vertical-align:top;
                          font-size:0.75rem;color:#6b7280;line-height:1.35;font-weight:400;">
                        {row['bestell_detail'] or '—'}
                      </td>
                      <td style="width:120px;padding:0.7rem 1.75rem 0.7rem 0;vertical-align:top;text-align:right;">
                        <div style="font-size:0.88rem;font-weight:600;color:#1a3a5c;
                            font-variant-numeric:tabular-nums;white-space:nowrap;line-height:1.2;">
                          CHF {row['ep_material']:,.2f}
                        </div>
                        <div style="font-size:0.60rem;color:#b0b8c4;white-space:nowrap;margin-top:2px;">/ Einheit</div>
                      </td>
                    </tr>"""

                # ── Gesamte System-Card als EIN st.markdown ──────────────────
                st.markdown(f"""<div style="background:white;border-radius:8px;margin-bottom:1.75rem;
                            box-shadow:0 1px 6px rgba(0,0,0,0.08);overflow:hidden;">
                  <div style="background:{col_accent};padding:1rem 1.75rem;
                               display:flex;align-items:center;justify-content:space-between;">
                    <div style="display:flex;align-items:center;gap:14px;">
                      <div style="background:rgba(255,255,255,0.18);border-radius:4px;
                                   padding:3px 10px;font-size:0.7rem;font-weight:700;
                                   letter-spacing:0.12em;color:white;font-family:'DM Mono',monospace;">
                        SYSTEM {sid}
                      </div>
                      <span style="font-size:0.95rem;font-weight:600;color:white;">
                        {n_schichten} Schichten{dritt_badge}
                      </span>
                    </div>
                    <div style="text-align:right;">
                      <div style="font-size:1.6rem;font-weight:700;color:white;line-height:1;
                          font-variant-numeric:tabular-nums;">
                        {sys_area:,.2f} m²
                      </div>
                      <div style="font-size:0.65rem;color:rgba(255,255,255,0.7);
                                   letter-spacing:0.08em;text-transform:uppercase;margin-top:2px;">
                        Systemfläche
                      </div>
                    </div>
                  </div>
                  <table style="width:100%;border-collapse:collapse;table-layout:auto;">
                    <thead>
                      <tr style="background:#f8f9fb;border-bottom:1.5px solid #eaecf0;">
                        <th style="padding:0.55rem 0 0.55rem 1.75rem;"></th>
                        <th style="padding:0.55rem 1rem 0.55rem 0;text-align:left;font-size:0.62rem;
                            font-weight:700;letter-spacing:0.09em;text-transform:uppercase;color:#9ca3af;">Schichtfunktion</th>
                        <th style="padding:0.55rem 1rem 0.55rem 0;text-align:left;font-size:0.62rem;
                            font-weight:700;letter-spacing:0.09em;text-transform:uppercase;color:#9ca3af;">Produkt / Art.-Nr.</th>
                        <th style="padding:0.55rem 1rem 0.55rem 0;text-align:right;font-size:0.62rem;
                            font-weight:700;letter-spacing:0.09em;text-transform:uppercase;color:#9ca3af;">Fläche</th>
                        <th style="padding:0.55rem 1rem 0.55rem 0;text-align:left;font-size:0.62rem;
                            font-weight:700;letter-spacing:0.09em;text-transform:uppercase;color:#9ca3af;">Bestellmenge</th>
                        <th style="padding:0.55rem 1.75rem 0.55rem 0;text-align:right;font-size:0.62rem;
                            font-weight:700;letter-spacing:0.09em;text-transform:uppercase;color:#9ca3af;">EP / Einheit</th>
                      </tr>
                    </thead>
                    <tbody>
                      {rows_html}
                    </tbody>
                  </table>
                </div>""", unsafe_allow_html=True)

            # ────────────────────────────────────────────────────────────────
            # ABSCHNITT B – KOSTENAUSWERTUNG
            # ────────────────────────────────────────────────────────────────
            st.markdown("""
            <div class="sw-section-bar" style="margin-top:3rem;">
              <div class="sw-section-bar-line" style="background:#e63946;"></div>
              <span class="sw-section-bar-title">B · Kostenauswertung nach System</span>
            </div>
            <p class="sw-section-bar-sub">
              Material- und Einbaukosten je System, CHF/m² Vergleich und Gesamtübersicht.
            </p>
            """, unsafe_allow_html=True)

            # Systemkosten aggregieren (inkl. Systemfläche korrekt)
            sys_records = []
            for idx, sid in enumerate(systems):
                sys_df = df_k[df_k["system_id"] == sid]
                base_rows = sys_df[sys_df["function"].str.startswith("01_")]
                sys_area = (base_rows["flaeche_m2"].sum()
                            if not base_rows.empty
                            else sys_df["flaeche_m2"].max())
                total   = sys_df["poskosten_total"].sum()
                mat     = sys_df["poskosten_mat"].sum()
                ein     = sys_df["poskosten_ein"].sum()
                n_pos   = len(sys_df)
                chf_m2  = total / sys_area if sys_area > 0 else 0
                sys_records.append({
                    "sid": sid, "area": sys_area, "mat": mat, "ein": ein,
                    "total": total, "chf_m2": chf_m2, "n_pos": n_pos,
                    "color": PALETTE[idx % len(PALETTE)]
                })

            grand_total = sum(r["total"] for r in sys_records)

            # System-Kosten-Cards — st.columns, 1 st.markdown pro Karte
            n_cols = min(len(sys_records), 2)
            cols_b = st.columns(n_cols)
            for bidx, rec in enumerate(sys_records):
                bpct = (rec["total"] / grand_total * 100) if grand_total > 0 else 0
                with cols_b[bidx % n_cols]:
                    st.markdown(f"""<div style="background:white;border-radius:8px;
                        margin-bottom:1rem;box-shadow:0 1px 5px rgba(0,0,0,0.07);
                        overflow:hidden;font-family:'Source Sans 3','Helvetica Neue',Arial,sans-serif;">
                      <div style="height:4px;background:{rec['color']};"></div>
                      <div style="padding:1.25rem 1.5rem;">
                        <div style="display:flex;justify-content:space-between;
                            align-items:flex-start;margin-bottom:0.75rem;">
                          <div>
                            <div style="font-size:0.62rem;font-weight:700;letter-spacing:0.12em;
                                text-transform:uppercase;color:#9ca3af;margin-bottom:4px;">
                              System {rec['sid']}
                            </div>
                            <div style="font-size:1.6rem;font-weight:600;color:#1a3a5c;
                                line-height:1;font-variant-numeric:tabular-nums;">
                              CHF {rec['total']:,.0f}
                            </div>
                          </div>
                          <div style="text-align:right;">
                            <div style="font-size:0.62rem;letter-spacing:0.08em;
                                text-transform:uppercase;color:#b0b8c4;margin-bottom:2px;">Anteil</div>
                            <div style="font-size:1.4rem;font-weight:600;color:{rec['color']};">
                              {bpct:.1f}%
                            </div>
                          </div>
                        </div>
                        <div style="background:#f0f2f5;border-radius:4px;height:4px;
                            margin-bottom:1rem;">
                          <div style="background:{rec['color']};height:4px;border-radius:4px;
                              width:{min(bpct,100):.1f}%;"></div>
                        </div>
                        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;
                            gap:0.5rem;margin-bottom:0.75rem;">
                          <div style="background:#f8f9fb;border-radius:5px;
                              padding:0.6rem 0.5rem;text-align:center;">
                            <div style="font-size:0.88rem;font-weight:600;color:#1a3a5c;
                                font-variant-numeric:tabular-nums;">
                              {rec['area']:,.1f}
                            </div>
                            <div style="font-size:0.58rem;color:#9ca3af;text-transform:uppercase;
                                letter-spacing:0.06em;margin-top:2px;">m²</div>
                          </div>
                          <div style="background:#f8f9fb;border-radius:5px;
                              padding:0.6rem 0.5rem;text-align:center;">
                            <div style="font-size:0.88rem;font-weight:600;color:{rec['color']};
                                font-variant-numeric:tabular-nums;">
                              {rec['chf_m2']:,.1f}
                            </div>
                            <div style="font-size:0.58rem;color:#9ca3af;text-transform:uppercase;
                                letter-spacing:0.06em;margin-top:2px;">CHF/m²</div>
                          </div>
                          <div style="background:#f8f9fb;border-radius:5px;
                              padding:0.6rem 0.5rem;text-align:center;">
                            <div style="font-size:0.88rem;font-weight:600;color:#374151;
                                font-variant-numeric:tabular-nums;">
                              {rec['n_pos']}
                            </div>
                            <div style="font-size:0.58rem;color:#9ca3af;text-transform:uppercase;
                                letter-spacing:0.06em;margin-top:2px;">Pos.</div>
                          </div>
                        </div>
                        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.5rem;
                            border-top:1px solid #f0f2f5;padding-top:0.6rem;">
                          <div style="padding:0.4rem 0.6rem;border-left:3px solid #e5e7eb;">
                            <div style="font-size:0.58rem;color:#b0b8c4;margin-bottom:3px;
                                text-transform:uppercase;letter-spacing:0.08em;">Material</div>
                            <div style="font-size:0.84rem;font-weight:500;color:#374151;
                                font-variant-numeric:tabular-nums;letter-spacing:0.01em;">
                              CHF {rec['mat']:,.0f}
                            </div>
                          </div>
                          <div style="padding:0.4rem 0.6rem;border-left:3px solid #e5e7eb;">
                            <div style="font-size:0.58rem;color:#b0b8c4;margin-bottom:3px;
                                text-transform:uppercase;letter-spacing:0.08em;">Einbau</div>
                            <div style="font-size:0.84rem;font-weight:500;color:#374151;
                                font-variant-numeric:tabular-nums;letter-spacing:0.01em;">
                              CHF {rec['ein']:,.0f}
                            </div>
                          </div>
                        </div>
                      </div>
                    </div>""", unsafe_allow_html=True)

            # Charts: CHF/m² Balken + Donut — untereinander, volle Breite
            st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)

            # ── Chart 1: CHF/m² nach System (Balken) ─────────────────────────
            st.markdown('<div class="sw-card">', unsafe_allow_html=True)
            st.markdown('<p class="sw-chart-title">CHF/m² nach System</p>', unsafe_allow_html=True)
            fig_m2 = go.Figure(go.Bar(
                x=[f"System {r['sid']}" for r in sys_records],
                y=[r["chf_m2"] for r in sys_records],
                marker_color=[r["color"] for r in sys_records],
                text=[f"CHF {r['chf_m2']:,.1f}/m²" for r in sys_records],
                textposition="outside",
                textfont=dict(size=13, family="Source Sans 3, Arial"),
                cliponaxis=False,
                hovertemplate="<b>%{x}</b><br>%{text}<extra></extra>",
                width=0.55,
            ))
            max_m2 = max((r["chf_m2"] for r in sys_records), default=0)
            fig_m2.update_layout(
                margin=dict(l=20, r=20, t=30, b=30),
                height=340,
                bargap=0.35,
                xaxis=dict(showgrid=False, title="",
                           tickfont=dict(size=12, color="#374151")),
                yaxis=dict(showgrid=True, gridcolor="#f0f2f5",
                           title="CHF/m²", tickformat=",.0f",
                           range=[0, max_m2 * 1.18]),
                plot_bgcolor="white",
                paper_bgcolor="white",
                font=dict(family="Source Sans 3, Arial", size=12, color="#1a1a1a"),
                showlegend=False,
            )
            st.plotly_chart(fig_m2, use_container_width=True, config={"displayModeBar": False})
            st.markdown('</div>', unsafe_allow_html=True)

            # ── Chart 2: Kostenverteilung (Donut) ────────────────────────────
            st.markdown('<div class="sw-card">', unsafe_allow_html=True)
            st.markdown('<p class="sw-chart-title">Kostenverteilung nach System</p>', unsafe_allow_html=True)
            fig_pie = go.Figure(go.Pie(
                labels=[f"System {r['sid']}" for r in sys_records],
                values=[r["total"] for r in sys_records],
                hole=0.58,
                marker_colors=[r["color"] for r in sys_records],
                marker=dict(line=dict(color="white", width=2)),
                textinfo="label+percent",
                textposition="outside",
                textfont=dict(size=12, family="Source Sans 3, Arial"),
                hovertemplate="<b>%{label}</b><br>CHF %{value:,.0f} (%{percent})<extra></extra>",
                sort=False,
            )) 
            fig_pie.update_layout(
                margin=dict(l=40, r=40, t=30, b=30),
                height=420,
                showlegend=False,
                plot_bgcolor="white",
                paper_bgcolor="white",
                font=dict(family="Source Sans 3, Arial", size=12, color="#1a1a1a"),
                annotations=[dict(
                    text=f"<b>CHF {grand_total:,.0f}</b><br><span style='font-size:11px;color:#9ca3af;'>Total exkl. MWST</span>",
                    x=0.5, y=0.5,
                    font=dict(size=16, family="Source Sans 3, Arial", color="#1a3a5c"),
                    showarrow=False
                )],
            )
            st.plotly_chart(fig_pie, use_container_width=True, config={"displayModeBar": False})
            st.markdown('</div>', unsafe_allow_html=True)

            # MWST-Totalbox
            mwst_rate = 0.081
            mwst_betrag = grand_total * mwst_rate
            total_inkl = grand_total * (1 + mwst_rate)
            st.markdown(f"""
            <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:1px;
                         background:#e0e0e0;border-radius:6px;overflow:hidden;margin:0.5rem 0 1.5rem;">
              <div style="background:white;padding:1rem 1.5rem;">
                <div style="font-size:0.65rem;font-weight:700;letter-spacing:0.1em;
                             text-transform:uppercase;color:#9ca3af;margin-bottom:4px;">
                  Total exkl. MWST
                </div>
                <div style="font-size:1.35rem;font-weight:600;color:#1a3a5c;
                             font-family:'Source Sans 3',Arial,sans-serif;
                             font-variant-numeric:tabular-nums;letter-spacing:0.01em;">
                  CHF {grand_total:,.2f}
                </div>
              </div>
              <div style="background:white;padding:1rem 1.5rem;">
                <div style="font-size:0.65rem;font-weight:700;letter-spacing:0.1em;
                             text-transform:uppercase;color:#9ca3af;margin-bottom:4px;">
                  MWST 8.1%
                </div>
                <div style="font-size:1.35rem;font-weight:600;color:#6b7280;
                             font-family:'Source Sans 3',Arial,sans-serif;
                             font-variant-numeric:tabular-nums;letter-spacing:0.01em;">
                  CHF {mwst_betrag:,.2f}
                </div>
              </div>
              <div style="background:#1a3a5c;padding:1rem 1.5rem;">
                <div style="font-size:0.65rem;font-weight:700;letter-spacing:0.1em;
                             text-transform:uppercase;color:rgba(255,255,255,0.6);margin-bottom:4px;">
                  Total inkl. MWST
                </div>
                <div style="font-size:1.35rem;font-weight:700;color:white;
                             font-family:'Source Sans 3',Arial,sans-serif;
                             font-variant-numeric:tabular-nums;letter-spacing:0.01em;">
                  CHF {total_inkl:,.2f}
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Detailtabelle
            st.markdown('<div class="sw-card">', unsafe_allow_html=True)
            st.markdown('<p class="sw-chart-title">Alle Positionen – Detailansicht</p>', unsafe_allow_html=True)

            df_table = df_k[[
                "system_id", "func_label", "product_name", "article_nr",
                "flaeche_m2", "bestell_detail",
                "ep_material", "ep_einbau",
                "poskosten_mat", "poskosten_ein", "poskosten_total",
                "is_drittprodukt"
            ]].copy()
            df_table.columns = [
                "System", "Funktion", "Produkt", "Art.-Nr.",
                "Fläche m²", "Bestellmenge",
                "EP Mat.", "EP Ein.",
                "Material CHF", "Einbau CHF", "Total CHF",
                "Drittprodukt"
            ]
            st.dataframe(
                df_table,
                hide_index=True,
                use_container_width=True,
                height=380,
                column_config={
                    "Total CHF":    st.column_config.NumberColumn(format="CHF %.2f"),
                    "Material CHF": st.column_config.NumberColumn(format="CHF %.2f"),
                    "Einbau CHF":   st.column_config.NumberColumn(format="CHF %.2f"),
                    "EP Mat.":      st.column_config.NumberColumn(format="CHF %.2f"),
                    "EP Ein.":      st.column_config.NumberColumn(format="CHF %.2f"),
                    "Fläche m²":    st.column_config.NumberColumn(format="%.2f m²"),
                }
            )
            st.markdown("""
            <div style="font-size:0.72rem;color:#9ca3af;margin-top:6px;">
              Drittprodukt = True: Platzhalter CHF 0.00 – Preis manuell erfassen.
            </div>
            """, unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # ── Protokoll & Download ──────────────────────────────────────────────
        st.markdown('<div class="sw-card">', unsafe_allow_html=True)
        st.markdown("""
        <div class="sw-section-header">
          <div class="sw-section-num" style="background:#555">4</div>
          <span class="sw-section-title">Protokoll &amp; Download</span>
        </div>
        """, unsafe_allow_html=True)
        st.markdown(f'<div class="sw-log">{results_k["log"]}</div>', unsafe_allow_html=True)
        st.markdown('<hr class="sw-divider">', unsafe_allow_html=True)
        if results_k["excel_bytes"]:
            st.download_button(
                "↓  Kostenauswertung_Swisspor.xlsx herunterladen",
                data=results_k["excel_bytes"],
                file_name="Kostenauswertung_Swisspor.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Keine Kostenpositionen — enriched IFC und Preisdatenbank prüfen.")
        st.markdown('</div>', unsafe_allow_html=True)

# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="sw-footer">
    swisspor AG · Bahnhofstrasse 50 · CH-6312 Steinhausen
    &nbsp;·&nbsp; swissporBIM v5.0 · DC_BAT_FS26 · HSLU
</div>
""", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
