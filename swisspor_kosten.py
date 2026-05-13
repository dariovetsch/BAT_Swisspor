"""
swisspor_kosten.py  –  BIM-ENGINE-SP  v4.0
============================================================
UC-02: Mengenermittlung & Kostenauswertung aus enriched IFC

v4.0 Änderungen:
  - Einzelschicht-first  (Gesamtaufbau als Fallback)
  - 4 Mengentypen: m², Liter, lfm/Rollen, m³+t, Stk
  - Verschnitt: +10 % Rollenware, +5 % Platten/Schüttgut
  - Einbaukosten (EP_Einbau) aus Preisdatenbank
  - Drittprodukte: CHF 0.00 Platzhalter aus Lookup_Drittprodukte
  - Produktname (IFC) als eigene Spalte
  - BKP-Code als Strukturierungsebene
  - 2-Sheet Excel: Kostenauswertung + BKP-Zusammenzug
============================================================
"""
import io, re, math
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import ifcopenshell


# ── Pset -> Funktionscode ──────────────────────────────────────────────────────
PSET_TO_FUNC = {
    "Pset_Swisspor_Produkt":                "01_Unterkonstruktion",
    "Pset_Swisspor_Haftvermittler":         "02_Haftvermittler",
    "Pset_Swisspor_Dampfbremse":            "03_Dampfbremse",
    "Pset_Swisspor_Dämmung":                "04_Wärmedämmung",
    "Pset_Swisspor_Gefälledämmung":         "05_Gefälledämmung",
    "Pset_Swisspor_Unterbahn":              "06_Unterbahn",
    "Pset_Swisspor_Oberbahn":              "07_Oberbahn",
    "Pset_Swisspor_Trennschicht":           "08_Trennschicht",
    "Pset_Swisspor_Drainageschicht":        "09_Drainageschicht",
    "Pset_Swisspor_Filterschicht":          "10_Filterschicht",
    "Pset_Swisspor_Wasserspeicherschicht":  "11_Wasserspeicherschicht",
    "Pset_Swisspor_Schutzschicht":          "12_Schutzschicht",
    "Pset_Swisspor_Splitschicht":           "13_Splitschicht",
    "Pset_Swisspor_Brandschutzschicht":     "14_Brandschutzschicht",
    "Pset_Swisspor_Vegetationstragschicht": "15_Vegetationstragschicht",
    "Pset_Swisspor_Belag":                  "16_Belag",
    "Pset_Swisspor_Absturzsicherung":       "17_Absturzsicherung",
}
_LAYER_PSETS = set(PSET_TO_FUNC.keys())

# ── Mengentypen ───────────────────────────────────────────────────────────────
ROLLEN_TYPES = {
    "Unterbahn", "Oberbahn", "Oberbahn wurzelecht",
    "Dampfbremse / Bauzeitabdichtung", "Trennschicht",
    "Drainageschicht", "Filterschicht",
    "Wasserspeicher- / Drainageelement", "Brandschutzschicht",
}
PLATTEN_TYPES  = {"Wärmedämmung", "Gefälledämmung"}
SCHUETT_TYPES  = {"Schutzschicht", "Splitschicht", "Vegetationstragschicht"}
STK_TYPES      = {"Absturzsicherung"}
HAFTV_TYPES    = {"Haftvermittler"}

# ── Verschnitt-Faktoren (auf Bestellmenge + Materialkosten) ──────────────────
_VFAK = {"rollen": 1.10, "platten": 1.05, "schuett": 1.05, "default": 1.00}

def _verschnitt(product_type: str) -> float:
    if product_type in ROLLEN_TYPES:  return _VFAK["rollen"]
    if product_type in PLATTEN_TYPES: return _VFAK["platten"]
    if product_type in SCHUETT_TYPES: return _VFAK["schuett"]
    return _VFAK["default"]


# ── Rollenlänge / -breite aus Format-String ───────────────────────────────────
def _parse_rollen(fmt: str):
    """'10 x 1.0 m Rolle' -> (laenge=10.0, breite=1.0)"""
    m = re.search(r"([\d.]+)\s*x\s*([\d.]+)", fmt or "")
    if not m:
        return None, None
    a, b = float(m.group(1)), float(m.group(2))
    return max(a, b), min(a, b)   # (länger, kürzer=Breite)


# ── Bestellmengen-Berechnung ──────────────────────────────────────────────────
def _bestell(rec: dict) -> dict:
    """
    Gibt zurück:
      menge_display  – angezeigte Bestellmenge
      einheit        – Einheit für Anzeige
      detail         – Kurztext für Excel-Zelle
      menge_material – Menge für Materialkostenberechnung (inkl. Verschnitt)
    """
    pt   = rec.get("product_type", "") or ""
    area = rec["area"]
    vfak = _verschnitt(pt)

    # Absturzsicherung als Sonderposition über Laufmeter ────────────────
    # Im Referenzmodell werden keine einzelnen Anschlagpunkte modelliert.
    # Darum wird die Länge / der Umfang als nachvollziehbare Mengenbasis verwendet.
    func = (rec.get("function", "") or "").lower()
    pname = (rec.get("product_name", "") or "").lower()
    if pt in STK_TYPES or "absturz" in func or "safsys" in pname:
        length_m = rec.get("length_m")
        if length_m is not None and length_m > 0:
            length_m = round(float(length_m), 2)
            return {"menge_display": length_m, "einheit": "lfm",
                    "detail": f"{length_m:.2f} lfm",
                    "menge_material": length_m,
                    "menge_einbau": length_m,
                    "mengenbasis": "Länge Begrenzungsrahmen"}

        n = rec.get("stk_count", 1)
        return {"menge_display": n, "einheit": "Stk",
                "detail": f"{n} Stk", "menge_material": n,
                "menge_einbau": n, "mengenbasis": "Anzahl Elemente"}

    # Liter (Haftvermittler) ─────────────────────────────────────────────
    if pt in HAFTV_TYPES:
        cons = float(rec.get("consumption_per_area") or 0.3)
        liter = round(area * cons, 1)
        return {"menge_display": liter, "einheit": "L",
                "detail": f"{liter:.1f} L  ({cons} L/m²)",
                "menge_material": area,   # priced per m²
                "menge_einbau": area,
                "mengenbasis": "IFC-Fläche"}

    # m³ + Tonnen (Schüttgut) ────────────────────────────────────────────
    if pt in SCHUETT_TYPES:
        thick_m = float(rec.get("min_thickness_mm") or 50) / 1000
        vol_net  = round(area * thick_m, 2)
        vol_best = round(vol_net * vfak, 2)
        detail = f"{vol_best:.2f} m³"
        if rec.get("density"):
            t = round(vol_best * float(rec["density"]) / 1000, 2)
            detail += f"  /  {t:.2f} t"
        if vfak > 1:
            detail += f"  (+{int((vfak-1)*100)}% Verschnitt)"
        return {"menge_display": vol_best, "einheit": "m³",
                "detail": detail, "menge_material": area * vfak,
                "menge_einbau": area * vfak,
                "mengenbasis": "IFC-Fläche"}

    # Laufmeter + Rollen ─────────────────────────────────────────────────
    if pt in ROLLEN_TYPES:
        laenge, breite = _parse_rollen(rec.get("format_str", ""))
        area_best = round(area * vfak, 2)
        if breite and breite > 0:
            lfm   = round(area_best / breite, 1)
            rollen = math.ceil(lfm / laenge) if laenge else "?"
            detail = (f"{lfm:.1f} lfm  ->  {rollen} Rollen "
                      f"({laenge}x{breite}m)  (+{int((vfak-1)*100)}% Verschnitt)")
            return {"menge_display": lfm, "einheit": "lfm",
                    "detail": detail, "menge_material": area_best,
                    "menge_einbau": area_best,
                    "mengenbasis": "IFC-Fläche"}
        return {"menge_display": area_best, "einheit": "m²",
                "detail": f"{area_best:.2f} m²  (+{int((vfak-1)*100)}%)",
                "menge_material": area_best,
                "menge_einbau": area_best,
                "mengenbasis": "IFC-Fläche"}

    # Platten (Dämmung, Gefälledämmung) ──────────────────────────────────
    if pt in PLATTEN_TYPES:
        area_best = round(area * vfak, 2)
        detail = (f"{area_best:.2f} m²  "
                  f"(+{int((vfak-1)*100)}% Verschnitt)")
        return {"menge_display": area_best, "einheit": "m²",
                "detail": detail, "menge_material": area_best,
                "menge_einbau": area_best,
                "mengenbasis": "IFC-Fläche"}

    # Standard m² ────────────────────────────────────────────────────────
    return {"menge_display": area, "einheit": "m²",
            "detail": f"{area:.2f} m²", "menge_material": area,
            "menge_einbau": area,
            "mengenbasis": "IFC-Fläche"}


# ── Preisdatenbank laden ──────────────────────────────────────────────────────
def load_preisdatenbank(xlsx_bytes: bytes):
    logs = []
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    # Haupt-Lookup (Art.-Nr. -> Preis + Einbau)
    lk = next((s for s in wb.sheetnames if "lookup_kosten" in s.lower()),
               next((s for s in wb.sheetnames if "lookup" in s.lower()),
                    wb.sheetnames[0]))
    art_db = {}
    for row in wb[lk].iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        art = str(row[0]).strip()
        try:    preis  = float(row[4]) if row[4] is not None else 0.0
        except: preis  = 0.0
        try:    einbau = float(row[5]) if len(row) > 5 and row[5] is not None else 0.0
        except: einbau = 0.0
        art_db[art] = {
            "bez":    str(row[1] or "").strip(),
            "preis":  preis,
            "einbau": einbau,
        }
    logs.append(f"OK Preisdatenbank: {len(art_db)} Artikel (Sheet '{lk}')")

    # Drittprodukte-Lookup (CostLabel -> Preis + Einbau)
    dritt_db = {}
    dk = next((s for s in wb.sheetnames if "drittprodukt" in s.lower()), None)
    if dk:
        for row in wb[dk].iter_rows(min_row=3, values_only=True):
            if not row[0]: continue
            label = str(row[0]).strip()
            try:    preis  = float(row[2]) if row[2] is not None else 0.0
            except: preis  = 0.0
            try:    einbau = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
            except: einbau = 0.0
            dritt_db[label] = {
                "bez":    str(row[1] or "").strip(),
                "preis":  preis,
                "einbau": einbau,
            }
        logs.append(f"OK Drittprodukte: {len(dritt_db)} Einträge (Sheet '{dk}')")
    else:
        logs.append("WARNUNG: Sheet 'Lookup_Drittprodukte' nicht gefunden – "
                    "Drittprodukte werden mit CHF 0.00 ausgewiesen")

    return art_db, dritt_db, logs


# ── Alle swisspor-Properties eines Elements einsammeln ───────────────────────
def _collect_sw_props(el) -> dict:
    """Alle Pset_Swisspor_*-Properties in ein flaches Dict."""
    props = {}
    for rel in el.IsDefinedBy:
        if not rel.is_a("IfcRelDefinesByProperties"): continue
        pdef = rel.RelatingPropertyDefinition
        if not pdef.is_a("IfcPropertySet"): continue
        if "swisspor" not in (pdef.Name or "").lower(): continue
        for p in pdef.HasProperties:
            if p.is_a("IfcPropertySingleValue") and p.NominalValue:
                props[p.Name] = str(p.NominalValue.wrappedValue).strip()
    return props


def _to_float(value):
    """Wandelt IFC-/Stringwerte robust in float um."""
    if value is None:
        return None
    if hasattr(value, "wrappedValue"):
        value = value.wrappedValue
    try:
        return float(value)
    except Exception:
        pass
    try:
        s = str(value).strip()
        s = (s.replace("m²", "").replace("m2", "")
               .replace("m³", "").replace("m3", "")
               .replace("lfm", "").replace("m", ""))
        s = s.replace("'", "").replace(" ", "").replace(",", ".")
        return float(s)
    except Exception:
        return None


def _get_area(el):
    """
    Beste IfcQuantityArea aus allen IfcElementQuantity des Elements.

    ArchiCAD exportiert viele Flächenwerte pro Bauteil, darunter Kanten-Flächen
    ("Kanten-Oberflächenbereich (netto)"), die NICHT die Bauteil-Grundfläche
    repräsentieren. Diese werden explizit ausgeschlossen.

    Priorität (höchste zuerst):
      1. "Oberflächenbereich oben (netto)"  (ArchiCAD Dach/Decke)
      2. "NetArea" / "NetFloorArea"          (IFC-Standard)
      3. "Oberflächenbereich oben"
      4. "Oberflächenbereich" / "GrossArea"
      5. Grösster verbleibender Nicht-Kanten-Wert
    """
    candidates = []
    for rel in getattr(el, "IsDefinedBy", []) or []:
        if not rel.is_a("IfcRelDefinesByProperties"):
            continue
        pdef = rel.RelatingPropertyDefinition
        if not pdef or not pdef.is_a("IfcElementQuantity"):
            continue
        for q in pdef.Quantities:
            if q.is_a("IfcQuantityArea"):
                v = _to_float(q.AreaValue)
                if v is not None and v > 0.1:
                    candidates.append((q.Name or "", v))

    if not candidates:
        return None

    name_lo = [(n.strip().lower(), v) for n, v in candidates]

    # Kanten-, Rand- und Öffnungs-Flächen ausschliessen
    _KANTEN_WORDS = ("kanten", "kantenfl", "rand", "edge", "opening",
                     "öffnung", "tür", "fenster", "löcher")
    non_kanten = [(n, v) for n, v in name_lo
                  if not any(kw in n for kw in _KANTEN_WORDS)]
    pool = non_kanten if non_kanten else name_lo

    # Prioritätsliste (exakt zuerst, dann Teilstring)
    preferred = [
        "oberflächenbereich oben (netto)",
        "nettofläche",
        "netfloorarea",
        "netarea",
        "net area",
        "grossarea",
        "grossfloorarea",
        "oberflächenbereich oben",
        "oberflächenbereich",
        "fläche",
    ]
    # Exakter Treffer
    for pref in preferred:
        for n, v in pool:
            if n == pref:
                return v
    # Teilstring-Treffer
    for pref in preferred:
        for n, v in pool:
            if pref in n:
                return v
    # Letzter Fallback: grösster Wert
    return max(v for _, v in pool)


def _get_length(el):
    """
    Liest eine Länge für Sonderpositionen wie Absturzsicherung.

    Priorität (höchste zuerst):
      1. "Länge Begrenzungsrahmen" (Swisspor-spezifisch)
      2. IFC-Standard "Length" / "Perimeter"
      3. ArchiCAD Wand-Referenzlinien-Länge (für SAFSYS-Wandelemente)
      4. "Grundriss-Umfang" / "Umfang" nur als letzter Fallback
         (ArchiCAD: Umfang = 2 x Länge + 2 x Dicke – für einzelne Wandsegmente
          daher ungeeignet als primäre Mengenbasis)
    """
    candidates = []
    for rel in getattr(el, "IsDefinedBy", []) or []:
        if not rel.is_a("IfcRelDefinesByProperties"):
            continue
        pdef = rel.RelatingPropertyDefinition
        if not pdef or not pdef.is_a("IfcElementQuantity"):
            continue
        for q in pdef.Quantities:
            if q.is_a("IfcQuantityLength"):
                v = _to_float(q.LengthValue)
                if v is not None and v > 0:
                    candidates.append((q.Name or "", v))

    if not candidates:
        return None

    preferred = [
        # Swisspor-spezifisch
        "länge begrenzungsrahmen",
        "laenge begrenzungsrahmen",
        # IFC-Standard
        "perimeter",
        "length",
        # ArchiCAD Wand-Längenquantitäten (SAFSYS-Wandelemente)
        "länge der referenzlinie",
        "laenge der referenzlinie",
        "wandlänge an der außenseite",
        "wandlänge an der innenseite",
        "durchschnittliche länge der wand",
        "3d-länge",
        # Allgemeine Teilsuche
        "laenge",
        "länge",
        # Umfang als letzter Fallback
        "umfang",
    ]
    for pref in preferred:
        for name, val in candidates:
            if str(name).strip().lower() == pref:
                return val
    for pref in preferred:
        for name, val in candidates:
            if pref in str(name).strip().lower():
                return val

    return candidates[0][1]


# ── IFC einlesen: Einzelschicht-first, Gesamtaufbau-Fallback ─────────────────
def extract_from_ifc(ifc_file_raw):
    logs       = []
    direct_recs = {}   # (sid, func, art, product, thickness, format) -> rec
    gesamt_recs = {}   # (sid, func, art, product, thickness, format) -> rec

    ifc = ifcopenshell.file.from_string(
        ifc_file_raw.getvalue().decode("utf-8", errors="replace")
    )
    elements = ifc.by_type("IfcElement")
    logs.append(f"OK IFC geladen: {len(elements)} Elemente")

    for el in elements:
        # ── Identifikation ──────────────────────────────────────────────
        sid = func = None
        for rel in el.IsDefinedBy:
            if not rel.is_a("IfcRelDefinesByProperties"): continue
            pdef = rel.RelatingPropertyDefinition
            if not (pdef.is_a("IfcPropertySet") and
                    pdef.Name == "Pset_Swisspor_Identifikation"):
                continue
            for p in pdef.HasProperties:
                if not (p.is_a("IfcPropertySingleValue") and p.NominalValue): continue
                pn = (p.Name or "").lower()
                pv = str(p.NominalValue.wrappedValue).strip()
                if "systemid" in pn:
                    try:    sid = str(int(float(pv))).zfill(2)
                    except: sid = pv.zfill(2) if pv else None
                if pn == "function":
                    func = pv
        if not (sid and func): continue

        area = _get_area(el)
        length_m = _get_length(el)

        # ════════════════════════════════════════════════════════════════
        # STRATEGIE 1 – Einzelschicht (func ≠ 00_Gesamtdachaufbau)
        # ════════════════════════════════════════════════════════════════
        if func != "00_Gesamtdachaufbau":
            if not area: continue
            props = _collect_sw_props(el)
            art = props.get("ArticleNumber") or props.get("Artikelnummer") or ""
            product_name = (props.get("ProductName") or props.get("Product") or "")
            min_thickness = props.get("MinThickness") or ""
            format_str = props.get("Format", "") or ""

            # Nicht nur nach System + Funktion gruppieren.
            # Im Modell kann dieselbe Funktion mehrfach vorkommen, z.B. zwei Wärmedämmungen
            # im gleichen System mit unterschiedlichen Flächen oder Teilflächen.
            key = (sid, func, art, product_name, min_thickness, format_str)

            base = {
                "article_nr":          art,
                "product_name":        product_name,
                "product_type":        props.get("ProductType", ""),
                "cost_code":           props.get("CostCode", ""),
                "cost_label":          props.get("CostLabel", ""),
                "price_unit":          props.get("PriceUnit", "m²"),
                "consumption_per_area":props.get("ConsumptionPerArea"),
                "min_thickness_mm":    min_thickness,
                "density":             props.get("Density"),
                "format_str":          format_str,
                "length_m":            length_m,
            }
            if key not in direct_recs:
                direct_recs[key] = {**base, "areas": [round(area, 3)],
                                    "lengths": [], "stk_count": 1}
                if length_m and length_m > 0:
                    direct_recs[key]["lengths"].append(round(float(length_m), 3))
            else:
                direct_recs[key]["areas"].append(round(area, 3))
                direct_recs[key]["stk_count"] += 1
                if length_m and length_m > 0:
                    direct_recs[key]["lengths"].append(round(float(length_m), 3))

        # ════════════════════════════════════════════════════════════════
        # STRATEGIE 2 – Gesamtaufbau (func == 00_Gesamtdachaufbau)
        # ════════════════════════════════════════════════════════════════
        else:
            if not area: continue
            layer_data = {}   # pset_name -> props-dict

            for rel in el.IsDefinedBy:
                if not rel.is_a("IfcRelDefinesByProperties"): continue
                pdef = rel.RelatingPropertyDefinition
                if not pdef.is_a("IfcPropertySet"): continue
                pname = pdef.Name or ""
                if pname not in _LAYER_PSETS: continue
                props = {}
                for p in pdef.HasProperties:
                    if p.is_a("IfcPropertySingleValue") and p.NominalValue:
                        props[p.Name] = str(p.NominalValue.wrappedValue).strip()
                if props.get("ArticleNumber") or True:   # include even w/o art
                    layer_data[pname] = props

            for pname, props in layer_data.items():
                f_code = PSET_TO_FUNC[pname]
                art = props.get("ArticleNumber", "") or ""
                product_name = (props.get("Product") or props.get("ProductName") or "")
                min_thickness = props.get("MinThickness") or ""
                format_str = props.get("Format", "") or ""

                # Gleiche Logik wie bei direkt modellierten Schichten:
                # Gruppierung nach System, Funktion, Artikel, Produkt, Dicke und Format.
                key = (sid, f_code, art, product_name, min_thickness, format_str)

                base = {
                    "article_nr":          art,
                    "product_name":        product_name,
                    "product_type":        props.get("ProductType", ""),
                    "cost_code":           props.get("CostCode", ""),
                    "cost_label":          props.get("CostLabel", ""),
                    "price_unit":          props.get("PriceUnit", "m²"),
                    "consumption_per_area":props.get("ConsumptionPerArea"),
                    "min_thickness_mm":    min_thickness,
                    "density":             props.get("Density"),
                    "format_str":          format_str,
                    "length_m":            length_m,
                }
                if key not in gesamt_recs:
                    gesamt_recs[key] = {**base, "areas": [round(area, 3)],
                                        "lengths": [], "stk_count": 1}
                    if length_m and length_m > 0:
                        gesamt_recs[key]["lengths"].append(round(float(length_m), 3))
                else:
                    gesamt_recs[key]["areas"].append(round(area, 3))
                    if length_m and length_m > 0:
                        gesamt_recs[key]["lengths"].append(round(float(length_m), 3))
                    if props.get("ProductType", "") in STK_TYPES:
                        gesamt_recs[key]["stk_count"] += 1

    # ── Merge: Einzelschicht priorisiert ─────────────────────────────────────
    records = []
    n_direkt = n_gesamt = 0

    direct_keys = set(direct_recs.keys())

    for key, rec in direct_recs.items():
        sid, func, art, product_name, min_thickness, format_str = key
        total_area   = round(sum(rec["areas"]), 3)
        lengths      = rec.get("lengths") or []
        total_length = round(sum(lengths), 3) if lengths else None
        records.append({
            **rec,
            "system_id":  sid,
            "function":   func,
            "area":       total_area,
            "length_m":   total_length,
            "mengenbasis": "Einzelschicht"
        })
        n_direkt += 1

    for key, rec in gesamt_recs.items():
        sid, func, art, product_name, min_thickness, format_str = key
        if key in direct_keys:
            continue   # exakt gleiche Einzelschicht hat Vorrang
        total_area   = round(sum(rec["areas"]), 3)
        lengths      = rec.get("lengths") or []
        total_length = round(sum(lengths), 3) if lengths else None
        records.append({
            **rec,
            "system_id":  sid,
            "function":   func,
            "area":       total_area,
            "length_m":   total_length,
            "mengenbasis": "Gesamtaufbau-Fallback"
        })
        n_gesamt += 1

    logs.append(f"\nINFO: {len(records)} Positionen extrahiert "
                f"({n_direkt} Einzelschicht, {n_gesamt} Gesamtaufbau-Fallback)")
    return records, logs


# ── Kostenberechnung ──────────────────────────────────────────────────────────
def calculate_costs(records, art_db: dict, dritt_db: dict):
    logs  = []
    kosten = []
    by_sys = defaultdict(list)
    for r in records:
        by_sys[r["system_id"]].append(r)

    for sid in sorted(by_sys):
        logs.append(f"\n─── System {sid} ───")
        for rec in sorted(by_sys[sid], key=lambda x: x["function"]):
            art        = rec["article_nr"]
            cost_label = rec.get("cost_label", "")
            bm         = _bestell(rec)

            # Preis-Lookup: Art.-Nr. -> Drittprodukt (CostLabel) -> 0.00
            if art and art in art_db:
                src    = art_db[art]
                is_dritt = False
            elif cost_label and cost_label in dritt_db:
                src    = dritt_db[cost_label]
                is_dritt = True
                logs.append(f"  INFO Drittprodukt: {cost_label}")
            else:
                src    = {"bez": rec.get("product_name", "—"), "preis": 0.0, "einbau": 0.0}
                is_dritt = True
                logs.append(f"  WARNUNG KEIN PREIS: {rec['function']} | {cost_label or art or '?'}")

            ep_mat  = src["preis"]
            ep_ein  = src.get("einbau", 0.0)
            menge_m = bm["menge_material"]

            pos_mat  = round(ep_mat  * menge_m, 2)
            menge_ein = bm.get("menge_einbau", rec["area"])
            pos_ein  = round(ep_ein  * menge_ein, 2)
            pos_tot  = round(pos_mat + pos_ein, 2)

            # Flaeche vs. Laenge trennen: Absturzsicherung wird ueber lfm abgerechnet,
            # nicht ueber m². flaeche_m2 = 0.0 verhindert falsche m²-Darstellung.
            is_absturz = (
                rec.get("product_type", "") in STK_TYPES
                or "absturz" in (rec.get("function", "") or "").lower()
                or "safsys" in (rec.get("product_name", "") or "").lower()
            )
            flaeche_out = 0.0 if is_absturz else rec["area"]

            kosten.append({
                "system_id":        sid,
                "function":         rec["function"],
                "cost_code":        rec.get("cost_code", "—"),
                "cost_label":       cost_label,
                "product_name":     rec.get("product_name", ""),
                "article_nr":       art,
                "bezeichnung":      src["bez"],
                "flaeche_m2":       flaeche_out,
                "laenge_m":         rec.get("length_m"),
                "mengenbasis":      bm.get("mengenbasis") or rec.get("mengenbasis", "IFC-Fläche"),
                "bestell_detail":   bm["detail"],
                "bestell_einheit":  bm["einheit"],
                "menge_material":   menge_m,
                "ep_material":      ep_mat,
                "ep_einbau":        ep_ein,
                "poskosten_mat":    pos_mat,
                "poskosten_ein":    pos_ein,
                "poskosten_total":  pos_tot,
                "is_drittprodukt":  is_dritt,
            })
            status = "WARNUNG" if is_dritt else "OK"
            logs.append(
                f"  {status} {rec['function']:28} | {art or cost_label:12}"
                f" | {bm['detail']:35}"
                f" | Mat CHF {pos_mat:>8,.2f}"
                f" | Ein CHF {pos_ein:>8,.2f}"
                f" | Tot CHF {pos_tot:>8,.2f}"
            )

    return kosten, logs


# ── Excel-Hilfsfunktionen ─────────────────────────────────────────────────────
NAVY = "1A3A5C"; RED = "E63946"; ORANGE = "F0A500"
LIGHT = "EEF2F7"; GREY = "F5F5F5"; WHITE = "FFFFFF"
GREEN = "E8F5E9"; YELLOW = "FFF9C4"

def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _st(c, bold=False, bg=WHITE, fg="1A1A1A", align="left",
        size=10, fmt=None):
    c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=False)
    c.border    = _bdr()
    if fmt: c.number_format = fmt


# ── Excel bauen ───────────────────────────────────────────────────────────────
def build_excel(kosten: list, mwst: float = 0.081) -> bytes:
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════
    # SHEET 1 – Kostenauswertung nach System
    # ════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Kostenauswertung"

    # Titel
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = "Swisspor – Kostenauswertung  |  Preisstand 2026  |  inkl. Verschnitt  |  exkl. MWST"
    _st(c, bold=True, bg=NAVY, fg=WHITE, align="center", size=11)
    ws.row_dimensions[1].height = 26

    # Header
    headers = ["Sys", "BKP", "Schichtfunktion", "Produktname (IFC)",
               "Art.-Nr.", "Fläche m²", "Bestellmenge",
               "EP Mat.", "EP Ein.", "Mat. CHF", "Ein. CHF", "Total CHF"]
    for i, h in enumerate(headers, 1):
        _st(ws.cell(2, i, h), bold=True, bg=NAVY, fg=WHITE, align="center", size=9)
    ws.row_dimensions[2].height = 20

    col_w = [6, 11, 24, 26, 12, 10, 32, 10, 10, 14, 14, 16]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rn = 3
    cur_sys = None
    sys_totals: dict = {}
    grand_mat = grand_ein = 0.0

    def _write_subtotal(row_n, sid, tot_mat, tot_ein):
        ws.merge_cells(f"A{row_n}:K{row_n}")
        _st(ws.cell(row_n, 1, f"  Subtotal System {sid}"),
            bold=True, bg=NAVY, fg=WHITE, align="left", size=9)
        _st(ws.cell(row_n, 12, tot_mat + tot_ein),
            bold=True, bg=NAVY, fg=WHITE, align="right", size=9,
            fmt='#,##0.00 "CHF"')
        ws.row_dimensions[row_n].height = 18

    for rec in kosten:
        sid = rec["system_id"]
        if sid != cur_sys:
            if cur_sys:
                _write_subtotal(rn, cur_sys,
                                sys_totals[cur_sys]["mat"],
                                sys_totals[cur_sys]["ein"])
                rn += 1
                for col in range(1, 13):
                    ws.cell(rn, col).fill = PatternFill("solid", fgColor=LIGHT)
                ws.row_dimensions[rn].height = 4
                rn += 1
            cur_sys = sid
            sys_totals[sid] = {"mat": 0.0, "ein": 0.0}

        bg = YELLOW if rec["is_drittprodukt"] else (GREY if rn % 2 == 0 else WHITE)
        sys_totals[sid]["mat"] += rec["poskosten_mat"]
        sys_totals[sid]["ein"] += rec["poskosten_ein"]
        grand_mat += rec["poskosten_mat"]
        grand_ein += rec["poskosten_ein"]

        vals  = [sid, rec["cost_code"], rec["function"],
                 rec["product_name"], rec["article_nr"],
                 rec["flaeche_m2"], rec["bestell_detail"],
                 rec["ep_material"], rec["ep_einbau"],
                 rec["poskosten_mat"], rec["poskosten_ein"], rec["poskosten_total"]]
        aligns = ["center","center","left","left","center",
                  "right","left","right","right","right","right","right"]
        fmts   = [None, None, None, None, None,
                  "#,##0.00", None,
                  "#,##0.00", "#,##0.00",
                  '#,##0.00', '#,##0.00', '#,##0.00 "CHF"']

        for i, (v, a, f) in enumerate(zip(vals, aligns, fmts), 1):
            cell_bg = bg
            if i == 12 and isinstance(v, (int, float)) and v > 0 and not rec["is_drittprodukt"]:
                cell_bg = GREEN
            _st(ws.cell(rn, i, v), bg=cell_bg, align=a, fmt=f, size=9)
        ws.row_dimensions[rn].height = 15
        rn += 1

    # Letzter Subtotal
    if cur_sys:
        _write_subtotal(rn, cur_sys,
                        sys_totals[cur_sys]["mat"],
                        sys_totals[cur_sys]["ein"])
        rn += 1

    # Totale
    rn += 1
    grand_tot = grand_mat + grand_ein
    for label, val, bg in [
        ("TOTAL Material exkl. MWST", grand_mat, LIGHT),
        ("TOTAL Einbau exkl. MWST",   grand_ein, LIGHT),
        ("TOTAL exkl. MWST",          grand_tot, NAVY),
        ("MWST 8.1 %",                round(grand_tot * mwst, 2), LIGHT),
        ("TOTAL inkl. MWST",          round(grand_tot * (1 + mwst), 2), RED),
    ]:
        ws.merge_cells(f"A{rn}:K{rn}")
        fg2 = WHITE if bg in (NAVY, RED) else "1A1A1A"
        _st(ws.cell(rn, 1, label),  bold=True, bg=bg, fg=fg2,
            align="right", size=11)
        _st(ws.cell(rn, 12, val),   bold=True, bg=bg, fg=fg2,
            align="right", size=11, fmt='#,##0.00 "CHF"')
        ws.row_dimensions[rn].height = 24
        rn += 1

    ws.freeze_panes = "A3"

    # ════════════════════════════════════════════════════════════════════
    # SHEET 2 – BKP-Zusammenzug
    # ════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("BKP-Zusammenzug")
    ws2.merge_cells("A1:G1")
    c2 = ws2["A1"]
    c2.value = "Swisspor – Kostenübersicht nach BKP-Code  |  alle Systeme summiert"
    _st(c2, bold=True, bg=NAVY, fg=WHITE, align="center", size=11)
    ws2.row_dimensions[1].height = 26

    bkp_headers = ["BKP-Code", "Bezeichnung", "Fläche m²",
                   "Material CHF", "Einbau CHF", "Total CHF", "Anteil %"]
    for i, h in enumerate(bkp_headers, 1):
        _st(ws2.cell(2, i, h), bold=True, bg=NAVY, fg=WHITE, align="center", size=9)
    ws2.row_dimensions[2].height = 20

    bkp_sums: dict = {}
    for rec in kosten:
        code = rec["cost_code"] or "—"
        label = rec["cost_label"] or rec["function"]
        if code not in bkp_sums:
            bkp_sums[code] = {"label": label, "flaeche": 0.0, "mat": 0.0, "ein": 0.0}
        bkp_sums[code]["flaeche"] += rec["flaeche_m2"] or 0.0
        bkp_sums[code]["mat"]     += rec["poskosten_mat"]
        bkp_sums[code]["ein"]     += rec["poskosten_ein"]

    r2 = 3
    for code in sorted(bkp_sums):
        s   = bkp_sums[code]
        tot = s["mat"] + s["ein"]
        pct = (tot / grand_tot * 100) if grand_tot > 0 else 0.0
        bg  = GREY if r2 % 2 == 0 else WHITE
        vals2 = [code, s["label"],
                 round(s["flaeche"], 2), s["mat"], s["ein"], tot, pct / 100]
        alns2 = ["center","left","right","right","right","right","right"]
        fms2  = [None, None, "#,##0.00",
                 '#,##0.00 "CHF"', '#,##0.00 "CHF"',
                 '#,##0.00 "CHF"', "0.0%"]
        for i, (v, a, f) in enumerate(zip(vals2, alns2, fms2), 1):
            cell_bg = GREEN if i == 6 and tot > 0 else bg
            _st(ws2.cell(r2, i, v), bg=cell_bg, align=a, fmt=f, size=9)
        ws2.row_dimensions[r2].height = 15
        r2 += 1

    # BKP Total
    r2 += 1
    ws2.merge_cells(f"A{r2}:E{r2}")
    _st(ws2.cell(r2, 1, "TOTAL exkl. MWST"),
        bold=True, bg=NAVY, fg=WHITE, align="right", size=11)
    _st(ws2.cell(r2, 6, grand_tot),
        bold=True, bg=NAVY, fg=WHITE, align="right", size=11,
        fmt='#,##0.00 "CHF"')
    _st(ws2.cell(r2, 7, 1.0),
        bold=True, bg=NAVY, fg=WHITE, align="right", size=11, fmt="0.0%")
    ws2.row_dimensions[r2].height = 24

    for i, w in enumerate([12, 28, 12, 16, 16, 16, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Hauptfunktion ─────────────────────────────────────────────────────────────
def process_kosten(ifc_file_raw, preisdb_file_raw):
    logs = ["─── 1. Preisdatenbank ───"]
    art_db, dritt_db, db_l = load_preisdatenbank(preisdb_file_raw.getvalue())
    logs.extend(db_l)

    logs.append("\n─── 2. IFC lesen ───")
    records, ifc_l = extract_from_ifc(ifc_file_raw)
    logs.extend(ifc_l)

    logs.append("\n─── 3. Kosten berechnen ───")
    kosten, k_l = calculate_costs(records, art_db, dritt_db)
    logs.extend(k_l)

    grand    = sum(k["poskosten_total"] for k in kosten)
    n_sw     = sum(1 for k in kosten if not k["is_drittprodukt"])
    n_dritt  = sum(1 for k in kosten if k["is_drittprodukt"])
    n_sys    = len({k["system_id"] for k in kosten})
    n_miss   = len([r for r in records
                    if not r["article_nr"] and
                    not r.get("cost_label") in dritt_db])

    if kosten:
        logs.append(f"\nTotal exkl. MWST:  CHF {grand:,.2f}")
        logs.append(f"   Total inkl. MWST:  CHF {grand * 1.081:,.2f}")
        logs.append(f"   davon Drittprodukte (CHF 0.00): {n_dritt} Positionen")
        excel_bytes = build_excel(kosten)
    else:
        logs.append("\nWARNUNG: Keine Positionen – enriched IFC und Preisdatenbank prüfen.")
        excel_bytes = None

    return {
        "excel_bytes":  excel_bytes,
        "grand_total":  grand,
        "n_matched":    n_sw,
        "n_drittprodukt": n_dritt,
        "n_missing":    n_miss,
        "n_systems":    n_sys,
        "kosten_list":  kosten,   # für Charts
        "log":          "\n".join(logs),
    }
